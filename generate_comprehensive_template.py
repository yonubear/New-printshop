#!/usr/bin/env python3
"""
Generate Comprehensive Excel Import Template

This script creates a comprehensive Excel template with multiple sheets for importing:
1. Paper Options
2. Finishing Options
3. Print Pricing
4. Saved Prices

Each sheet includes example data and follows the database schema structure.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_header_style(sheet, row_num=1):
    """Apply header styling to the first row of the given sheet"""
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

def auto_adjust_columns(sheet):
    """Auto-adjust column widths based on content"""
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = min(adjusted_width, 50)

def add_info_header(sheet, title, instructions):
    """Add an information header to the sheet"""
    sheet.insert_rows(1, 3)
    
    # Title row
    sheet.cell(row=1, column=1, value=title)
    sheet.cell(row=1, column=1).font = Font(size=14, bold=True, color="1F4E78")
    
    # Instructions row
    sheet.cell(row=2, column=1, value=instructions)
    sheet.cell(row=2, column=1).font = Font(italic=True)
    
    # Merge cells across the width of the data
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=sheet.max_column)
    
    # Add an empty row for spacing
    sheet.row_dimensions[3].height = 10

def create_paper_options_sheet():
    """Create sample data for paper options"""
    data = {
        'name': [
            'Standard Copy Paper',
            'Premium Cardstock', 
            'Glossy Text',
            'Matte Cover',
            'Recycled Bond'
        ],
        'description': [
            '20# White Bond Paper',
            '110# Cover, White',
            '100# Gloss Text, White',
            '80# Matte Cover, White',
            '24# Recycled Bond, Natural White'
        ],
        'category': [
            'Bond',
            'Cover',
            'Text',
            'Cover',
            'Bond'
        ],
        'weight': [
            '20#',
            '110#',
            '100#',
            '80#',
            '24#'
        ],
        'size': [
            'Letter',
            'Letter',
            'Tabloid',
            'Letter',
            'Letter'
        ],
        'color': [
            'White',
            'White',
            'White',
            'White',
            'Natural White'
        ],
        'price_per_sheet': [
            0.02,
            0.12,
            0.08,
            0.10,
            0.03
        ],
    }
    
    return pd.DataFrame(data)

def create_finishing_options_sheet():
    """Create sample data for finishing options"""
    data = {
        'name': [
            'Lamination - 5mil Glossy',
            'Lamination - 10mil Matte',
            'Spiral Binding - Black',
            'Perfect Binding',
            'Corner Staple'
        ],
        'description': [
            '5mil Glossy Lamination for durability',
            '10mil Matte Lamination, non-glare finish',
            'Black Spiral Binding up to 1" diameter',
            'Perfect binding with square spine',
            'Single corner staple'
        ],
        'category': [
            'Lamination',
            'Lamination',
            'Binding',
            'Binding',
            'Stapling'
        ],
        'base_price': [
            2.00,
            3.00,
            5.00,
            15.00,
            0.50
        ],
        'price_per_piece': [
            0.25,
            0.35,
            0.00,
            0.00,
            0.00
        ],
        'price_per_sqft': [
            2.00,
            3.00,
            0.00,
            0.00,
            0.00
        ],
        'minimum_price': [
            5.00,
            7.50,
            5.00,
            15.00,
            0.50
        ]
    }
    
    return pd.DataFrame(data)

def create_print_pricing_sheet():
    """Create sample data for print pricing"""
    data = {
        'name': [
            'B&W Letter 1-sided',
            'B&W Letter 2-sided',
            'Color Letter 1-sided',
            'Color Letter 2-sided',
            'Color Tabloid 1-sided'
        ],
        'paper_size': [
            'Letter',
            'Letter',
            'Letter',
            'Letter',
            'Tabloid'
        ],
        'color_type': [
            'Black & White',
            'Black & White',
            'Full Color',
            'Full Color',
            'Full Color'
        ],
        'price_per_side': [
            0.10,
            0.08,
            0.50,
            0.45,
            0.90
        ],
        'cost_per_side': [
            0.03,
            0.03,
            0.15,
            0.15,
            0.30
        ],
        'notes': [
            'Standard B&W print on letter paper, single sided',
            'Standard B&W print on letter paper, double sided',
            'Full color print on letter paper, single sided',
            'Full color print on letter paper, double sided',
            'Full color print on tabloid paper, single sided'
        ]
    }
    
    return pd.DataFrame(data)

def create_saved_prices_sheet():
    """Create sample data for saved prices"""
    data = {
        'name': [
            'Business Cards - Standard',
            'Business Cards - Premium',
            'Flyer - Letter Size',
            'Brochure - Tri-Fold',
            'Booklet - Saddle Stitch'
        ],
        'description': [
            '3.5x2 inch, Full Color, 110# Gloss Cover',
            '3.5x2 inch, Full Color, Soft-Touch Lamination',
            '8.5x11 inch, Full Color, 100# Gloss Text',
            '8.5x11 inch folded to 3.7x8.5, Full Color both sides',
            '8.5x11 inch saddle stitched, 8 pages + cover'
        ],
        'sku': [
            'BC-STD',
            'BC-PRM',
            'FLY-LTR',
            'BRO-TRI',
            'BKL-SAD'
        ],
        'category': [
            'print_job',
            'print_job',
            'print_job',
            'print_job',
            'print_job'
        ],
        'cost_price': [
            5.00,
            8.50,
            0.35,
            0.75,
            3.50
        ],
        'price': [
            25.00,
            35.00,
            0.75,
            1.50,
            12.00
        ],
        'unit': [
            'pack',
            'pack',
            'each',
            'each',
            'each'
        ],
        'is_template': [
            True,
            True,
            True,
            True,
            True
        ],
        'materials': [
            'Premium Cardstock:25:sheets, Ink:2.5:ml, Cutting:1:service',
            'Premium Cardstock:25:sheets, Ink:2.5:ml, Lamination - 5mil Glossy:2:sqft, Cutting:1:service',
            'Glossy Text:1:sheet, Ink:1:ml',
            'Glossy Text:1:sheet, Ink:2:ml, Folding:1:service',
            'Glossy Text:4:sheets, Premium Cardstock:1:sheets, Ink:10:ml, Stapling:1:service'
        ]
    }
    
    return pd.DataFrame(data)

def create_comprehensive_excel():
    """Create a comprehensive Excel template with multiple sheets"""
    # Create dataframes for each sheet
    paper_df = create_paper_options_sheet()
    finishing_df = create_finishing_options_sheet()
    printing_df = create_print_pricing_sheet()
    saved_prices_df = create_saved_prices_sheet()
    
    # Create an Excel file with multiple sheets
    with pd.ExcelWriter('comprehensive_import_template.xlsx', engine='openpyxl') as writer:
        # Write each dataframe to a separate sheet
        paper_df.to_excel(writer, sheet_name='Paper Options', index=False)
        finishing_df.to_excel(writer, sheet_name='Finishing Options', index=False)
        printing_df.to_excel(writer, sheet_name='Print Pricing', index=False)
        saved_prices_df.to_excel(writer, sheet_name='Saved Prices', index=False)
        
        # Get workbook and apply formatting
        workbook = writer.book
        
        # Format Paper Options sheet
        paper_sheet = workbook['Paper Options']
        create_header_style(paper_sheet)
        auto_adjust_columns(paper_sheet)
        add_info_header(
            paper_sheet, 
            "Paper Options Import Template",
            "Add paper types with their specifications. These will be available for selection in order and quote forms."
        )
        
        # Format Finishing Options sheet
        finishing_sheet = workbook['Finishing Options']
        create_header_style(finishing_sheet)
        auto_adjust_columns(finishing_sheet)
        add_info_header(
            finishing_sheet, 
            "Finishing Options Import Template",
            "Add finishing options like lamination, binding, etc. These will be available as add-ons for orders and quotes."
        )
        
        # Format Print Pricing sheet
        printing_sheet = workbook['Print Pricing']
        create_header_style(printing_sheet)
        auto_adjust_columns(printing_sheet)
        add_info_header(
            printing_sheet, 
            "Print Pricing Import Template",
            "Add pricing configurations for different print options (per side). These will be used for automatic price calculations."
        )
        
        # Format Saved Prices sheet
        saved_prices_sheet = workbook['Saved Prices']
        create_header_style(saved_prices_sheet)
        auto_adjust_columns(saved_prices_sheet)
        add_info_header(
            saved_prices_sheet, 
            "Saved Prices Import Template",
            "Add common print jobs, materials, and services with their prices. For materials, use format: 'Name:Quantity:Unit, Name2:Quantity2:Unit2'"
        )
    
    print("Comprehensive Excel template created: comprehensive_import_template.xlsx")

if __name__ == "__main__":
    import sys
    
    # Check if output path is specified as argument
    if len(sys.argv) > 1:
        output_path = sys.argv[1]
        
        # Create workbook with pandas
        # Create dataframes for each sheet
        paper_df = create_paper_options_sheet()
        finishing_df = create_finishing_options_sheet()
        printing_df = create_print_pricing_sheet()
        saved_prices_df = create_saved_prices_sheet()
        
        # Create an Excel writer with the specified output path
        import pandas as pd
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write each dataframe to a separate sheet
            paper_df.to_excel(writer, sheet_name='Paper Options', index=False)
            finishing_df.to_excel(writer, sheet_name='Finishing Options', index=False)
            printing_df.to_excel(writer, sheet_name='Print Pricing', index=False)
            saved_prices_df.to_excel(writer, sheet_name='Saved Prices', index=False)
            
            # Get workbook and apply formatting
            workbook = writer.book
            
            # Format Paper Options sheet
            paper_sheet = workbook['Paper Options']
            create_header_style(paper_sheet)
            auto_adjust_columns(paper_sheet)
            add_info_header(
                paper_sheet, 
                "Paper Options Import Template",
                "Add paper types with their specifications. These will be available for selection in order and quote forms."
            )
            
            # Format Finishing Options sheet
            finishing_sheet = workbook['Finishing Options']
            create_header_style(finishing_sheet)
            auto_adjust_columns(finishing_sheet)
            add_info_header(
                finishing_sheet, 
                "Finishing Options Import Template",
                "Add finishing options like lamination, binding, etc. These will be available as add-ons for orders and quotes."
            )
            
            # Format Print Pricing sheet
            printing_sheet = workbook['Print Pricing']
            create_header_style(printing_sheet)
            auto_adjust_columns(printing_sheet)
            add_info_header(
                printing_sheet, 
                "Print Pricing Import Template",
                "Add pricing configurations for different print options (per side). These will be used for automatic price calculations."
            )
            
            # Format Saved Prices sheet
            saved_prices_sheet = workbook['Saved Prices']
            create_header_style(saved_prices_sheet)
            auto_adjust_columns(saved_prices_sheet)
            add_info_header(
                saved_prices_sheet, 
                "Saved Prices Import Template",
                "Add common print jobs, materials, and services with their prices. For materials, use format: 'Name:Quantity:Unit, Name2:Quantity2:Unit2'"
            )
        
        print(f"Comprehensive Excel template created: {output_path}")
    else:
        # Use default output path
        create_comprehensive_excel()